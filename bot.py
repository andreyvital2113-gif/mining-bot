# -*- coding: utf-8 -*-
"""
Telegram-бот "Дашборд производственных показателей"
Библиотека: pyTelegramBotAPI (telebot)
"""

import telebot
from telebot import types
import os
import re
import json
import io
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

TOKEN = os.environ.get("BOT_TOKEN", "ВСТАВЬ_СЮДА_ТОКЕН")
ADMIN_ID = os.environ.get("ADMIN_ID")  # твой Telegram user_id, только он может грузить Excel
bot = telebot.TeleBot(TOKEN)

UNIT = "м³"

MONTHS = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
          "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

MONTH_EMOJI = ["❄️", "❄️", "🌱", "🌱", "🌿", "☀️",
               "☀️", "☀️", "🍂", "🍂", "🍂", "❄️"]

IND = {
    "plan": "План по ОП",
    "stat": "Факт по статистике",
    "mark": "Факт марк. замеру",
    "uchet": "Факт принято к учету",
}
IND_TEXT_TO_KEY = {v: k for k, v in IND.items()}

IND_LABEL_TO_KEY = {
    "📊 Факт по статистике": "stat",
    "📏 Факт марк. замеру": "mark",
    "✅ Факт принято к учету": "uchet",
}

CATEGORY_ORDER = [
    "Горная масса",
    "Добыча",
    "Вскрыша и навалы",
    "Прочие",
    "Переработка (промывка) песков",
]

CATEGORY_EMOJI = {
    "Горная масса": "🪨",
    "Добыча": "⛏️",
    "Вскрыша и навалы": "🚛",
    "Прочие": "📦",
    "Переработка (промывка) песков": "💧",
}

NO_FULL = {"Переработка (промывка) песков"}  # нет марк.замера и учёта

# ====== ДАННЫЕ ПО ГОДАМ ======
# YEARS_DATA[год]["categories"][категория][показатель] = [12 значений]
# YEARS_DATA[год]["excavators"][экскаватор][показатель] = [12 значений]
YEARS_DATA = {
    "2026": {
        "categories": {
            "Горная масса": {
                "plan":  [73, 131, 78, 152, 0, 0, 0, 0, 0, 0, 0, 0],
                "stat":  [64, 88, 99, 115, 0, 0, 0, 0, 0, 0, 0, 0],
                "mark":  [63, 89, 99, 115, 0, 0, 0, 0, 0, 0, 0, 0],
                "uchet": [64, 89, 99, 115, 0, 0, 0, 0, 0, 0, 0, 0],
            },
            "Добыча": {
                "plan":  [0, 0, 0, 9, 0, 0, 0, 0, 0, 0, 0, 0],
                "stat":  [0, 0, 0, 20, 0, 0, 0, 0, 0, 0, 0, 0],
                "mark":  [0, 0, 0, 20, 0, 0, 0, 0, 0, 0, 0, 0],
                "uchet": [0, 0, 0, 20, 0, 0, 0, 0, 0, 0, 0, 0],
            },
            "Вскрыша и навалы": {
                "plan":  [52, 50, 27, 77, 0, 0, 0, 0, 0, 0, 0, 0],
                "stat":  [24, 55, 37, 65, 0, 0, 0, 0, 0, 0, 0, 0],
                "mark":  [23, 56, 37, 65, 0, 0, 0, 0, 0, 0, 0, 0],
                "uchet": [24, 56, 37, 65, 0, 0, 0, 0, 0, 0, 0, 0],
            },
            "Прочие": {
                "plan":  [21, 81, 51, 66, 0, 0, 0, 0, 0, 0, 0, 0],
                "stat":  [40, 33, 62, 30, 0, 0, 0, 0, 0, 0, 0, 0],
                "mark":  [40, 33, 62, 30, 0, 0, 0, 0, 0, 0, 0, 0],
                "uchet": [40, 33, 62, 30, 0, 0, 0, 0, 0, 0, 0, 0],
            },
            "Переработка (промывка) песков": {
                "plan":  [11, 10, 0, 9, 0, 0, 0, 0, 0, 0, 0, 0],
                "stat":  [5, 10, 0, 10, 0, 0, 0, 0, 0, 0, 0, 0],
                "mark":  [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                "uchet": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            },
        },
        "excavators": {},
    }
}

# ====== состояние пользователя ======
# chat_id -> {"year":.., "section": "categories"/"excavators", "item":.., "ind":.., "awaiting_period": bool}
user_state = {}


# ================= сохранение / загрузка =================
def load_saved_data():
    global YEARS_DATA
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
            YEARS_DATA.update(saved)
            print("Загружены сохранённые данные из data.json")
        except Exception as e:
            print(f"Не удалось загрузить data.json: {e}")


def save_data():
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(YEARS_DATA, f, ensure_ascii=False, indent=2)


# ================= парсинг Excel =================
def find_header_row(rows):
    """Находит строку с названиями показателей, возвращает (idx, [(col, ind_key), ...])."""
    for r_idx, row in enumerate(rows):
        matches = [(c_idx, IND_TEXT_TO_KEY[val]) for c_idx, val in enumerate(row)
                   if isinstance(val, str) and val.strip() in IND_TEXT_TO_KEY]
        if len(matches) >= 4:
            return r_idx, matches
    return None, []


def month_index_in_cell(row, max_cols=3):
    for cell in row[:max_cols]:
        if isinstance(cell, str):
            low = cell.strip().lower()
            for m_i, m_name in enumerate(MONTHS):
                if m_name.lower() in low:
                    return m_i
    return None


def parse_fixed_groups(rows, header_row_idx, header_cols, group_order):
    col_to_group = {}
    for i, (col_idx, ind_key) in enumerate(header_cols):
        g_idx = i // 4
        if g_idx < len(group_order):
            col_to_group[col_idx] = (group_order[g_idx], ind_key)

    result = {g: {k: [0] * 12 for k in IND} for g in group_order}
    months_found = 0
    for row in rows[header_row_idx + 1:]:
        m_idx = month_index_in_cell(row)
        if m_idx is None:
            continue
        months_found += 1
        for col_idx, (g, ind_key) in col_to_group.items():
            if col_idx < len(row):
                val = row[col_idx]
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (TypeError, ValueError):
                    val = 0
                result[g][ind_key][m_idx] = val
    return result, months_found


def parse_dynamic_groups(rows, header_row_idx, header_cols):
    """Для листа с экскаваторами: имена групп берутся из строки над заголовком."""
    label_row = rows[header_row_idx - 1] if header_row_idx > 0 else []

    group_names = []
    col_to_group = {}
    for i, (col_idx, ind_key) in enumerate(header_cols):
        g_idx = i // 4
        block_start_col = header_cols[g_idx * 4][0]
        if g_idx >= len(group_names):
            name = None
            for c in range(block_start_col, block_start_col + 4):
                if c < len(label_row) and isinstance(label_row[c], str) and label_row[c].strip():
                    name = label_row[c].strip()
                    break
            group_names.append(name or f"Экскаватор {g_idx + 1}")
        col_to_group[col_idx] = (group_names[g_idx], ind_key)

    result = {g: {k: [0] * 12 for k in IND} for g in group_names}
    months_found = 0
    for row in rows[header_row_idx + 1:]:
        m_idx = month_index_in_cell(row)
        if m_idx is None:
            continue
        months_found += 1
        for col_idx, (g, ind_key) in col_to_group.items():
            if col_idx < len(row):
                val = row[col_idx]
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (TypeError, ValueError):
                    val = 0
                result[g][ind_key][m_idx] = val
    return result, months_found


def find_single_indicator_columns(header_row):
    """Возвращает {ind_key: col_idx}, беря первое вхождение каждого показателя (не блоки по 4)."""
    result = {}
    for c_idx, val in enumerate(header_row):
        if isinstance(val, str) and val.strip() in IND_TEXT_TO_KEY:
            key = IND_TEXT_TO_KEY[val.strip()]
            if key not in result:
                result[key] = c_idx
    return result


def find_col_by_header_text(header_row, text):
    for c_idx, val in enumerate(header_row):
        if isinstance(val, str) and val.strip().lower() == text.lower():
            return c_idx
    return None


def parse_month_sheet_excavators(rows):
    """
    Разбирает лист конкретного месяца: ищет блок между строкой 'Экскавация'
    и строкой 'Переработка (промывка) песков', извлекает по каждому экскаватору
    строку 'Итого горная масса'.
    """
    header_row_idx = None
    name_col = work_col = None
    ind_cols = {}
    for r_idx, row in enumerate(rows):
        n_col = find_col_by_header_text(row, "Наименование")
        w_col = find_col_by_header_text(row, "Вид работ")
        inds = find_single_indicator_columns(row)
        if n_col is not None and w_col is not None and len(inds) >= 4:
            header_row_idx, name_col, work_col, ind_cols = r_idx, n_col, w_col, inds
            break
    if header_row_idx is None:
        return {}

    excavators = {}
    current_name = None
    for row in rows[header_row_idx + 1:]:
        name_val = row[name_col] if name_col < len(row) else None
        work_val = row[work_col] if work_col < len(row) else None
        name_text = name_val.strip() if isinstance(name_val, str) else None
        work_text = work_val.strip() if isinstance(work_val, str) else None

        if name_text:
            low = name_text.lower()
            if "переработка" in low:
                break
            if "экскавация" in low or "итого" in low:
                continue
            current_name = name_text

        if current_name and work_text and "итого горная масса" in work_text.lower():
            values = {}
            for ind_key, col in ind_cols.items():
                v = row[col] if col < len(row) else None
                try:
                    v = float(v)
                    if v == int(v):
                        v = int(v)
                except (TypeError, ValueError):
                    v = 0
                values[ind_key] = v
            excavators[current_name] = values

    return excavators


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- лист категорий ---
    sheet = wb["Сводная"] if "Сводная" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header_row_idx, header_cols = find_header_row(rows)
    if header_row_idx is None:
        raise ValueError("Не нашёл строку с показателями на листе «Сводная».")
    categories, months_found = parse_fixed_groups(rows, header_row_idx, header_cols, CATEGORY_ORDER)

    # --- экскаваторы: ищем по листам с названиями месяцев ---
    excavators = {}
    for m_idx, m_name in enumerate(MONTHS):
        sheet_name = None
        for s in wb.sheetnames:
            if s.strip().lower() == m_name.lower():
                sheet_name = s
                break
        if not sheet_name:
            continue
        m_rows = list(wb[sheet_name].iter_rows(values_only=True))
        month_result = parse_month_sheet_excavators(m_rows)
        for exc_name, vals in month_result.items():
            if exc_name not in excavators:
                excavators[exc_name] = {k: [0] * 12 for k in IND}
            for ind_key, v in vals.items():
                excavators[exc_name][ind_key][m_idx] = v

    return categories, excavators, months_found, len(excavators) > 0


def extract_year_from_filename(filename):
    match = re.search(r"(20\d{2})", filename)
    return match.group(1) if match else None


def fmt_num(v):
    """Форматирует число с разделением групп разрядов пробелом: 62000 -> 62 000."""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, int):
        return f"{v:,}".replace(",", " ")
    return f"{v:,.2f}".replace(",", " ")


# ================= форматирование отчёта =================
def format_report(year, section_label, item_name, item_emoji, ind_key_used, plan, fact,
                   month_range, fallback_note=""):
    lines = [f"{item_emoji} {item_name} · {year} год", f"Сравнение: План по ОП vs {IND[ind_key_used]}", ""]
    total_plan = total_fact = 0
    has_rows = False
    for i in month_range:
        p, f = plan[i], fact[i]
        if p == 0 and f == 0:
            continue
        has_rows = True
        total_plan += p
        total_fact += f
        pct = round(f / p * 100, 1) if p > 0 else None
        pct_str = f"{pct}%" if pct is not None else "—"
        mark = "✅" if pct and pct >= 100 else ("⚠️" if pct is not None else "")
        lines.append(f"{MONTH_EMOJI[i]} {MONTHS[i]}: план {fmt_num(p)} {UNIT} / факт {fmt_num(f)} {UNIT}  ({pct_str}) {mark}")

    if not has_rows:
        lines.append("Нет данных за выбранный период.")
    else:
        total_pct = round(total_fact / total_plan * 100, 1) if total_plan > 0 else None
        lines.append("")
        if total_pct is not None:
            lines.append(f"📈 Итого: план {fmt_num(total_plan)} {UNIT} / факт {fmt_num(total_fact)} {UNIT} ({total_pct}%)")
        else:
            lines.append(f"📈 Итого: план {fmt_num(total_plan)} {UNIT} / факт {fmt_num(total_fact)} {UNIT}")

    if fallback_note:
        lines.append(fallback_note)

    return "\n".join(lines)


def generate_chart(item_name, ind_label, plan, fact, month_range):
    months_sel, plan_vals, fact_vals = [], [], []
    for i in month_range:
        if plan[i] == 0 and fact[i] == 0:
            continue
        months_sel.append(MONTHS[i])
        plan_vals.append(plan[i])
        fact_vals.append(fact[i])

    if not months_sel:
        return None

    x = range(len(months_sel))
    width = 0.35
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar([i - width / 2 for i in x], plan_vals, width, label="План", color="#4C6EF5")
    ax.bar([i + width / 2 for i in x], fact_vals, width, label="Факт", color="#37B24D")
    ax.set_xticks(list(x))
    ax.set_xticklabels(months_sel, rotation=30, ha="right")
    ax.set_ylabel(f"Объём, {UNIT}")
    ax.set_title(f"{item_name}\n{ind_label}")
    ax.legend()
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


# ================= клавиатуры =================
def main_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🚀 Начать")
    return kb


def years_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    years = sorted(YEARS_DATA.keys())
    kb.add(*[f"📅 {y}" for y in years])
    return kb


def section_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    kb.add("📁 Общие показатели")
    kb.add("🚜 Показатели по экскаваторам")
    kb.add("📈 График выполнения")
    kb.add("⬅️ Назад")
    return kb


def cats_menu(year):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    for c in YEARS_DATA[year]["categories"]:
        kb.add(f"{CATEGORY_EMOJI.get(c, '📁')} {c}")
    kb.add("⬅️ Назад")
    return kb


def excavators_menu(year):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    exc = YEARS_DATA[year].get("excavators", {})
    if not exc:
        kb.add("Нет данных по экскаваторам")
    for e in exc:
        kb.add(f"🚜 {e}")
    kb.add("⬅️ Назад")
    return kb


def ind_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    kb.add("📊 Факт по статистике")
    kb.add("📏 Факт марк. замеру")
    kb.add("✅ Факт принято к учету")
    kb.add("⬅️ Назад")
    return kb


def period_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    kb.add(*[f"{MONTH_EMOJI[i]} {m}" for i, m in enumerate(MONTHS)])
    kb.add("🗓 Весь год")
    kb.add("✏️ Свой период")
    kb.add("⬅️ Назад")
    return kb


# ================= вспомогательное =================
def strip_emoji_prefix(text):
    return re.sub(r"^\S+\s+", "", text).strip()


def parse_custom_period(text):
    """Ищет два названия месяцев в свободном тексте, возвращает (start_idx, end_idx) или None."""
    low = text.lower()
    found = []
    for i, m in enumerate(MONTHS):
        if m.lower() in low:
            found.append(i)
    if not found:
        return None
    if len(found) == 1:
        return found[0], found[0]
    return min(found), max(found)


# ================= хендлеры =================
@bot.message_handler(commands=["myid"])
def my_id(msg):
    bot.send_message(msg.chat.id, f"🆔 Твой Telegram ID: `{msg.from_user.id}`", parse_mode="Markdown")


@bot.message_handler(commands=["start"])
def start(msg):
    user_state[msg.chat.id] = {"want_chart": False}
    bot.send_message(
        msg.chat.id,
        "📈 Статистика производственных показателей\n\n📅 Выбери год:",
        reply_markup=years_menu(),
    )


@bot.message_handler(func=lambda m: m.text.startswith("📅 ") and m.text.replace("📅 ", "") in YEARS_DATA)
def choose_section(msg):
    m_year = msg.text.replace("📅 ", "")
    user_state[msg.chat.id] = {"year": m_year, "want_chart": False}
    bot.send_message(msg.chat.id, "Выбери раздел:", reply_markup=section_menu())


@bot.message_handler(func=lambda m: m.text == "📈 График выполнения")
def enable_chart_mode(msg):
    st = user_state.setdefault(msg.chat.id, {})
    if "year" not in st:
        bot.send_message(msg.chat.id, "Сначала выбери год.", reply_markup=years_menu())
        return
    st["want_chart"] = True
    bot.send_message(
        msg.chat.id,
        "📈 Режим графика включён.\nВыбери, по чему построить график:",
        reply_markup=section_menu(),
    )


@bot.message_handler(func=lambda m: m.text == "📁 Общие показатели")
def open_categories(msg):
    st = user_state.setdefault(msg.chat.id, {})
    if "year" not in st:
        bot.send_message(msg.chat.id, "Сначала выбери год.", reply_markup=years_menu())
        return
    st["section"] = "categories"
    bot.send_message(msg.chat.id, "Выбери направление:", reply_markup=cats_menu(st["year"]))


@bot.message_handler(func=lambda m: m.text == "🚜 Показатели по экскаваторам")
def open_excavators(msg):
    st = user_state.setdefault(msg.chat.id, {})
    if "year" not in st:
        bot.send_message(msg.chat.id, "Сначала выбери год.", reply_markup=years_menu())
        return
    st["section"] = "excavators"
    bot.send_message(msg.chat.id, "Выбери экскаватор:", reply_markup=excavators_menu(st["year"]))


@bot.message_handler(func=lambda m: user_state.get(m.chat.id, {}).get("section") == "categories"
                      and strip_emoji_prefix(m.text) in YEARS_DATA.get(user_state.get(m.chat.id, {}).get("year", ""), {}).get("categories", {}))
def choose_item_cat(msg):
    st = user_state[msg.chat.id]
    st["item"] = strip_emoji_prefix(msg.text)
    bot.send_message(msg.chat.id, "С чем сравнить план?", reply_markup=ind_menu())


@bot.message_handler(func=lambda m: user_state.get(m.chat.id, {}).get("section") == "excavators"
                      and strip_emoji_prefix(m.text) in YEARS_DATA.get(user_state.get(m.chat.id, {}).get("year", ""), {}).get("excavators", {}))
def choose_item_exc(msg):
    st = user_state[msg.chat.id]
    st["item"] = strip_emoji_prefix(msg.text)
    bot.send_message(msg.chat.id, "С чем сравнить план?", reply_markup=ind_menu())


@bot.message_handler(func=lambda m: m.text in IND_LABEL_TO_KEY)
def choose_period(msg):
    st = user_state.setdefault(msg.chat.id, {})
    if "item" not in st:
        bot.send_message(msg.chat.id, "Сначала выбери направление/экскаватор.")
        return
    st["ind"] = IND_LABEL_TO_KEY[msg.text]
    st["awaiting_period"] = False
    bot.send_message(msg.chat.id, "За какой период?", reply_markup=period_menu())


@bot.message_handler(func=lambda m: m.text == "✏️ Свой период")
def ask_custom_period(msg):
    st = user_state.setdefault(msg.chat.id, {})
    if "ind" not in st:
        bot.send_message(msg.chat.id, "Сначала выбери показатель для сравнения.")
        return
    st["awaiting_period"] = True
    bot.send_message(
        msg.chat.id,
        "✏️ Напиши период текстом, например: Март-Сентябрь или Январь-Декабрь",
        reply_markup=types.ReplyKeyboardRemove(),
    )


def render_report(msg, month_range, period_label):
    st = user_state.get(msg.chat.id, {})
    year = st.get("year")
    section = st.get("section")
    item = st.get("item")
    ind_key = st.get("ind")

    if not all([year, section, item, ind_key]):
        bot.send_message(msg.chat.id, "Начни заново: /start")
        return

    data_block = YEARS_DATA[year]["categories" if section == "categories" else "excavators"]
    data = data_block[item]

    ind_key_used = ind_key
    fallback_note = ""
    if section == "categories" and item in NO_FULL and ind_key in ("mark", "uchet"):
        ind_key_used = "stat"
        fallback_note = "⚠️ Для этого направления нет марк.замера/учёта — показана статистика."

    plan = data["plan"]
    fact = data[ind_key_used]
    emoji = CATEGORY_EMOJI.get(item, "🚜") if section == "categories" else "🚜"

    report = format_report(year, section, item, emoji, ind_key_used, plan, fact, month_range, fallback_note)
    kb = cats_menu(year) if section == "categories" else excavators_menu(year)
    bot.send_message(msg.chat.id, report, reply_markup=kb)

    if st.get("want_chart"):
        chart_buf = generate_chart(item, IND[ind_key_used], plan, fact, month_range)
        if chart_buf:
            bot.send_photo(msg.chat.id, chart_buf)
        else:
            bot.send_message(msg.chat.id, "Нет данных для графика за выбранный период.")


@bot.message_handler(func=lambda m: m.text == "🗓 Весь год")
def preset_full_year(msg):
    render_report(msg, range(12), "Весь год")


@bot.message_handler(func=lambda m: strip_emoji_prefix(m.text) in MONTHS)
def single_month(msg):
    month_name = strip_emoji_prefix(msg.text)
    idx = MONTHS.index(month_name)
    render_report(msg, range(idx, idx + 1), month_name)


@bot.message_handler(func=lambda m: user_state.get(m.chat.id, {}).get("awaiting_period"))
def custom_period_text(msg):
    st = user_state[msg.chat.id]
    st["awaiting_period"] = False
    parsed = parse_custom_period(msg.text)
    if parsed is None:
        bot.send_message(msg.chat.id, "Не распознал период. Напиши, например: Март-Сентябрь", reply_markup=period_menu())
        return
    start_idx, end_idx = parsed
    render_report(msg, range(start_idx, end_idx + 1), msg.text)


@bot.message_handler(content_types=["document"])
def handle_excel(msg):
    if ADMIN_ID and str(msg.from_user.id) != str(ADMIN_ID):
        bot.send_message(msg.chat.id, "⛔ Загружать данные может только администратор.")
        return

    fname = msg.document.file_name or ""
    if not fname.lower().endswith((".xlsx", ".xlsm")):
        bot.send_message(msg.chat.id, "Пришли файл в формате .xlsx (Excel).")
        return

    year = extract_year_from_filename(fname)
    if not year:
        bot.send_message(
            msg.chat.id,
            "⚠️ Не нашёл год в названии файла. Переименуй файл так, чтобы он содержал год, "
            "например «Показатели_2026.xlsx», и пришли снова.",
        )
        return

    bot.send_message(msg.chat.id, f"📥 Получил файл за {year} год, разбираю...")
    try:
        file_info = bot.get_file(msg.document.file_id)
        downloaded = bot.download_file(file_info.file_path)
        tmp_path = f"/tmp/{msg.document.file_id}.xlsx"
        with open(tmp_path, "wb") as f:
            f.write(downloaded)

        categories, excavators, months_found, has_exc_sheet = parse_excel(tmp_path)
        os.remove(tmp_path)

        YEARS_DATA[year] = {
            "categories": categories,
            "excavators": excavators if has_exc_sheet else YEARS_DATA.get(year, {}).get("excavators", {}),
        }
        save_data()

        exc_note = f"\n🚜 Экскаваторов найдено: {len(excavators)}" if has_exc_sheet else \
            "\n🚜 Данные по экскаваторам не найдены (нет листов по месяцам с блоком «Экскавация» — «Переработка (промывка) песков»)."

        bot.send_message(
            msg.chat.id,
            f"✅ Данные за {year} год обновлены.\n"
            f"📁 Категорий: {len(CATEGORY_ORDER)}\n"
            f"📆 Месяцев с данными: {months_found}"
            f"{exc_note}\n\n📅 Выбери год для отчёта:",
            reply_markup=years_menu(),
        )
    except ValueError as e:
        bot.send_message(msg.chat.id, f"⚠️ Не смог разобрать файл: {e}")
    except Exception as e:
        bot.send_message(msg.chat.id, f"⚠️ Ошибка при обработке файла: {e}")


@bot.message_handler(func=lambda m: m.text == "⬅️ Назад")
def go_back(msg):
    st = user_state.get(msg.chat.id, {})
    st["awaiting_period"] = False
    if "ind" in st:
        st.pop("ind")
        section = st.get("section")
        item = st.get("item")
        if item:
            st.pop("item")
            bot.send_message(msg.chat.id, "Выбери направление:",
                              reply_markup=cats_menu(st["year"]) if section == "categories" else excavators_menu(st["year"]))
            return
    if "section" in st:
        st.pop("section")
        st.pop("item", None)
        st["want_chart"] = False
        bot.send_message(msg.chat.id, "Выбери раздел:", reply_markup=section_menu())
        return
    if "year" in st:
        st.pop("year")
        bot.send_message(msg.chat.id, "📅 Выбери год:", reply_markup=years_menu())
        return
    start(msg)


@bot.message_handler(func=lambda m: True)
def fallback(msg):
    bot.send_message(msg.chat.id, "🤔 Не понял. Выбери год:", reply_markup=years_menu())


def setup_commands():
    bot.set_my_commands([
        types.BotCommand("start", "🚀 Начать"),
    ])


if __name__ == "__main__":
    load_saved_data()
    setup_commands()
    print("Бот запущен...")
    bot.infinity_polling()
